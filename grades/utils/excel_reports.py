import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import os
from django.conf import settings
from django.contrib.auth.models import User
from ..models import Curso, Subject, Matricula, CursoActividad, Nota, Comportamiento, _trunc2

# --- ESTILOS COMPARTIDOS ---
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
VERTICAL_ALIGN = Alignment(textRotation=90, horizontal='center', vertical='center')
BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Paleta: Azul primario, Rosa secundario, Lavanda terciario
FILL_SUBJECTS = PatternFill(start_color="2C5DA7", end_color="2C5DA7", fill_type="solid")      # Azul menos intenso - cabeceras de materias
FILL_AVG = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")            # Azul claro - promedios
FILL_INSUMOS = PatternFill(start_color="E8DEF8", end_color="E8DEF8", fill_type="solid")        # Lavanda - insumos
FILL_TOTAL = PatternFill(start_color="E91E8C", end_color="E91E8C", fill_type="solid")          # Rosa - totales
FILL_TRIMESTER = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")      # Rosa claro - promedios trimestre
FILL_PROM_TRIM = PatternFill(start_color="EDE7F6", end_color="EDE7F6", fill_type="solid")      # Lavanda suave - prom. trimestral
FILL_GRAY = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")           # Azul claro - boletines
FILL_LIGHT_GRAY = PatternFill(start_color="EBF0FA", end_color="EBF0FA", fill_type="solid")     # Azul muy claro - cabecera boletín

def _insertar_cabecera_institucional(ws, total_cols, titulo, docente_nombre, curso_full, is_resumido=False):
    """Inserta la cabecera institucional estándar en cualquier reporte Excel."""
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 28
    ws.row_dimensions[4].height = 13
    ws.row_dimensions[5].height = 24
    ws.row_dimensions[6].height = 20
    
    mid = total_cols // 2
    # Ajuste de posición de logos basado en el ancho
    factor = 4 if is_resumido else 10
    col_logo1 = max(1, mid - factor)
    col_logo2 = (total_cols - 2) if is_resumido else min(total_cols, mid + (factor - 3))
    col_doc = max(1, mid - (factor + 2))

    try:
        path_min = os.path.join(settings.BASE_DIR, 'static', 'logos', 'misniterio.jpeg')
        path_esc = os.path.join(settings.BASE_DIR, 'static', 'logos', 'logoescuela.jpeg')
        if os.path.exists(path_min):
            img_min = Image(path_min); img_min.height = 75; img_min.width = 170
            ws.add_image(img_min, get_column_letter(col_logo1) + '2')
        if os.path.exists(path_esc):
            img_esc = Image(path_esc); img_esc.height = 90; img_esc.width = 100
            ws.add_image(img_esc, get_column_letter(col_logo2) + '2')
    except: pass

    # Filas de texto
    for r, val, sz in [(2, "UNIDAD EDUCATIVA", 13), (3, '“ARQUÉOLOGO JULIO VITERI GAMBOA”', 16), 
                       (4, "DIR: CDLA. CHIRIJOS AV. AMAZONA Y ERNESTO SEMINARIO", 9), (5, titulo.upper(), 12)]:
        ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=total_cols)
        c = ws.cell(row=r, column=1, value=val)
        c.font = Font(bold=True, size=sz); c.alignment = CENTER_ALIGN

    # Docente y Curso
    ws.cell(row=6, column=col_doc, value=f"DOCENTE: {docente_nombre.upper()}").font = Font(bold=True, size=11, underline="single")
    ws.merge_cells(start_row=6, end_row=6, start_column=max(1, col_doc + 4), end_column=total_cols)
    c_meta = ws.cell(row=6, column=max(1, col_doc + 4), value=curso_full)
    c_meta.font = Font(bold=True, size=11, underline="single"); c_meta.alignment = CENTER_ALIGN

def generar_excel_trimestre(curso_id, trimestre, resumido=False):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = f"Trimestre {trimestre}"
    curso = Curso.objects.get(pk=curso_id)
    subjects = curso.subjects.all().order_by('nombre')
    matriculas = Matricula.objects.filter(curso=curso).select_related('estudiante')

    docente_nombre = f"{curso.docente.first_name} {curso.docente.last_name}" if curso.docente else "ADMINISTRADOR"
    
    # --- OPTIMIZACIÓN DE CABECERA ---
    # Traemos todos los CursoActividad de una vez para no hacer queries en el bucle de columnas
    all_ca_trim = list(CursoActividad.objects.filter(curso=curso, trimestre=trimestre).select_related('actividad'))
    ca_by_subject = {}
    for ca in all_ca_trim:
        ca_by_subject.setdefault(ca.subject_id, []).append(ca)

    if resumido:
        total_cols = 2 + subjects.count() + 3
    else:
        total_cols = 2 + sum(len(ca_by_subject.get(s.id, [])) + 1 for s in subjects) + 3

    trim_label = "PRIMER" if trimestre == 1 else "SEGUNDO" if trimestre == 2 else "TERCER"
    curso_full = f"{str(curso.nivel).upper()}   {str(curso.periodo)}"
    _insertar_cabecera_institucional(ws, total_cols, f"REGISTRO DE CALIFICACIONES DEL {trim_label} TRIMESTRE", docente_nombre, curso_full, resumido)

    # Cabeceras
    start_row = 8
    ws.row_dimensions[start_row].height = 110 if resumido else 25
    ws.row_dimensions[start_row+1].height = 85
    
    for c, v in [(1, "N°"), (2, "ESTUDIANTES")]:
        cell = ws.cell(row=start_row, column=c, value=v)
        cell.font=HEADER_FONT; cell.fill=FILL_SUBJECTS; cell.alignment=CENTER_ALIGN; cell.border=BORDER
        ws.merge_cells(start_row=start_row, end_row=start_row+1, start_column=c, end_column=c)

    curr_col = 3
    subject_map = []
    for s in subjects:
        acts = [] if resumido else ca_by_subject.get(s.id, [])
        n_cols = len(acts) + 1
        
        cell = ws.cell(row=start_row, column=curr_col, value=s.nombre.upper())
        cell.font=HEADER_FONT; cell.fill=FILL_SUBJECTS; cell.border=BORDER
        if resumido: cell.alignment=VERTICAL_ALIGN
        else:
            cell.alignment=CENTER_ALIGN
            ws.merge_cells(start_row=start_row, end_row=start_row, start_column=curr_col, end_column=curr_col + n_cols - 1)
        
        if not resumido:
            for i, a in enumerate(acts):
                sc = ws.cell(row=start_row+1, column=curr_col + i, value=a.actividad.nombre.upper())
                sc.font=Font(bold=True, size=10); sc.fill=FILL_INSUMOS; sc.alignment=VERTICAL_ALIGN; sc.border=BORDER
        
        ac = ws.cell(row=start_row+1, column=curr_col + n_cols - 1, value="PROMEDIO")
        ac.font=Font(bold=True, size=10); ac.fill=FILL_AVG; ac.alignment=VERTICAL_ALIGN; ac.border=BORDER
        
        subject_map.append({'s': s, 'acts': acts, 'col': curr_col, 'n': n_cols})
        curr_col += n_cols

    final_headers = ["COMPORTAMIENTO", "TOTAL", f"{trimestre}° TRIMESTRE"]
    for i, h in enumerate(final_headers):
        cell = ws.cell(row=start_row, column=curr_col + i, value=h.upper())
        if i == 2: cell.font=Font(bold=True, size=10); cell.fill=FILL_TRIMESTER
        else: cell.font=HEADER_FONT; cell.fill=FILL_SUBJECTS
        cell.alignment=VERTICAL_ALIGN; cell.border=BORDER
        ws.merge_cells(start_row=start_row, end_row=start_row+1, start_column=curr_col + i, end_column=curr_col + i)

    # --- OPTIMIZACIÓN DE DATOS (NOTAS Y COMPORTAMIENTOS) ---
    mat_ids = [m.pk for m in matriculas]
    todas_notas = Nota.objects.filter(matricula_id__in=mat_ids, trimestre=trimestre)
    notas_map = {(n.matricula_id, n.curso_actividad_id): float(n.valor) for n in todas_notas if n.curso_actividad_id}
    # Para promedios por materia
    notas_by_subject = {}
    for n in todas_notas:
        key = (n.matricula_id, n.subject_id)
        notas_by_subject.setdefault(key, []).append(float(n.valor))

    comportamientos = {c.matricula_id: c.valor for c in Comportamiento.objects.filter(matricula_id__in=mat_ids, trimestre=trimestre)}

    # Datos
    r_num = start_row + 2
    for idx, mat in enumerate(matriculas, 1):
        ws.row_dimensions[r_num].height = 25
        ws.cell(row=r_num, column=1, value=idx).alignment = CENTER_ALIGN
        ws.cell(row=r_num, column=2, value=f"{mat.estudiante.apellidos} {mat.estudiante.nombres}").font = Font(bold=True, size=13)
        
        # --- REPLICAR LÓGICA DE MATRICULA MODEL ---
        subject_averages = []
        # Identificar materias activas (según logic de Matricula.get_trimestre_average)
        active_subject_ids = [s.id for s in subjects if len(ca_by_subject.get(s.id, [])) > 0]
        
        for sm in subject_map:
            sid = sm['s'].id
            if not resumido:
                for i, ca in enumerate(sm['acts']):
                    v = notas_map.get((mat.pk, ca.pk), 0.0)
                    c = ws.cell(row=r_num, column=sm['col'] + i, value=v)
                    c.number_format = '0.00'; c.alignment = CENTER_ALIGN; c.font = Font(size=12)
            
            # Cálculo exacto de get_subject_average
            vals = notas_by_subject.get((mat.pk, sid), [])
            count = len(ca_by_subject.get(sid, []))
            p = _trunc2(sum(vals) / count) if count > 0 else 0.0
            
            # Guardamos el promedio si la materia es activa para el total/promedio final
            if sid in active_subject_ids:
                subject_averages.append(p)
                
            cp = ws.cell(row=r_num, column=sm['col'] + sm['n'] - 1, value=p)
            cp.font=Font(bold=True, size=12); cp.fill=FILL_AVG; cp.number_format='0.00'; cp.alignment=CENTER_ALIGN

        # Comportamiento
        ws.cell(row=r_num, column=curr_col, value=comportamientos.get(mat.pk, 'B')).alignment=CENTER_ALIGN
        
        # Cálculo exacto de get_trimestre_total
        total_trim = _trunc2(sum(subject_averages))
        ws.cell(row=r_num, column=curr_col+1, value=float(total_trim)).number_format='0.00'
        
        # Cálculo exacto de get_trimestre_average
        prom_trim = _trunc2(total_trim / len(active_subject_ids)) if active_subject_ids else 0.0
        cf = ws.cell(row=r_num, column=curr_col+2, value=float(prom_trim))
        cf.font=Font(bold=True, size=12); cf.fill=FILL_TRIMESTER; cf.number_format='0.00'; cf.alignment=CENTER_ALIGN
        r_num += 1

    for r in range(start_row, r_num):
        for c in range(1, curr_col + 3): ws.cell(row=r, column=c).border = BORDER
    ws.column_dimensions['A'].width = 6; ws.column_dimensions['B'].width = 46
    for c in range(3, curr_col + 3): ws.column_dimensions[get_column_letter(c)].width = 7.5
    return wb

def generar_excel_anual(curso_id):
    """Genera el reporte consolidado anual (Cuadro Final)."""
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Consolidado Anual"
    curso = Curso.objects.get(pk=curso_id)
    subjects = curso.subjects.all().order_by('nombre')
    matriculas = Matricula.objects.filter(curso=curso).select_related('estudiante')
    docente_nombre = f"{curso.docente.first_name} {curso.docente.last_name}" if curso.docente else "ADMINISTRADOR"
    
    # 2 (N, Est) + 4 por materia (T1, T2, T3, Promedio) + 3 (Comp, Total, Prom Final)
    total_cols = 2 + (subjects.count() * 4) + 3
    curso_full = f"{str(curso.nivel).upper()}   {str(curso.periodo)}"
    _insertar_cabecera_institucional(ws, total_cols, "REGISTRO DE CALIFICACIONES FINAL", docente_nombre, curso_full)

    start_row = 8
    ws.row_dimensions[start_row].height = 25
    ws.row_dimensions[start_row+1].height = 95
    
    for c, v in [(1, "N°"), (2, "ESTUDIANTES")]:
        cell = ws.cell(row=start_row, column=c, value=v)
        cell.font=HEADER_FONT; cell.fill=FILL_SUBJECTS; cell.alignment=CENTER_ALIGN; cell.border=BORDER
        ws.merge_cells(start_row=start_row, end_row=start_row+1, start_column=c, end_column=c)

    curr_col = 3
    sub_headers = ["1ER TRIM.", "2DO TRIM.", "3ER TRIM.", "PROMEDIO"]
    for s in subjects:
        cell = ws.cell(row=start_row, column=curr_col, value=s.nombre.upper())
        cell.font=HEADER_FONT; cell.fill=FILL_SUBJECTS; cell.alignment=CENTER_ALIGN; cell.border=BORDER
        ws.merge_cells(start_row=start_row, end_row=start_row, start_column=curr_col, end_column=curr_col + 3)
        
        for i, h in enumerate(sub_headers):
            sc = ws.cell(row=start_row+1, column=curr_col + i, value=h)
            sc.font=Font(bold=True, size=10); sc.alignment=VERTICAL_ALIGN; sc.border=BORDER
            if i == 3: sc.fill = FILL_AVG
        curr_col += 4

    for i, h in enumerate(["COMPORT.", "TOTAL", "PROM. FINAL"]):
        cell = ws.cell(row=start_row, column=curr_col + i, value=h)
        if i == 2: cell.font=Font(bold=True, size=10); cell.fill=FILL_TRIMESTER
        else: cell.font=HEADER_FONT; cell.fill=FILL_SUBJECTS
        cell.alignment=VERTICAL_ALIGN; cell.border=BORDER
        ws.merge_cells(start_row=start_row, end_row=start_row+1, start_column=curr_col + i, end_column=curr_col + i)

    r_num = start_row + 2
    for idx, mat in enumerate(matriculas, 1):
        ws.row_dimensions[r_num].height = 25
        ws.cell(row=r_num, column=1, value=idx).alignment = CENTER_ALIGN
        ws.cell(row=r_num, column=2, value=f"{mat.estudiante.apellidos} {mat.estudiante.nombres}").font = Font(bold=True, size=13)
        
        sc_col = 3
        suma_finales = 0
        for s in subjects:
            t1 = mat.get_subject_average(s, 1)
            t2 = mat.get_subject_average(s, 2)
            t3 = mat.get_subject_average(s, 3)
            pf = _trunc2((t1 + t2 + t3) / 3)
            suma_finales += pf
            
            for i, v in enumerate([t1, t2, t3, pf]):
                c = ws.cell(row=r_num, column=sc_col + i, value=v)
                c.number_format = '0.00'; c.alignment = CENTER_ALIGN
                if i == 3: c.font = Font(bold=True, size=12); c.fill = FILL_AVG
            sc_col += 4

        comp = Comportamiento.objects.filter(matricula=mat, trimestre=4).first()
        ws.cell(row=r_num, column=curr_col, value=comp.valor if comp else 'B').alignment=CENTER_ALIGN
        
        # Total y Promedio Final (Exactamente como el sistema)
        ws.cell(row=r_num, column=curr_col+1, value=float(_trunc2(suma_finales))).number_format='0.00'
        
        prom_final = _trunc2(suma_finales / subjects.count()) if subjects.count() > 0 else 0.0
        cf = ws.cell(row=r_num, column=curr_col+2, value=float(prom_final))
        cf.font=Font(bold=True, size=12); cf.fill=FILL_TRIMESTER; cf.number_format='0.00'; cf.alignment=CENTER_ALIGN
        r_num += 1

    for r in range(start_row, r_num):
        for c in range(1, curr_col + 3): ws.cell(row=r, column=c).border = BORDER
    ws.column_dimensions['A'].width = 6; ws.column_dimensions['B'].width = 46
    for c in range(3, curr_col + 3): ws.column_dimensions[get_column_letter(c)].width = 7.5
    return wb

def generar_excel_boletines_individuales(curso_id):
    """Genera reportes de calificaciones individuales por estudiante (Boletines)."""
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Boletines Individuales"
    curso = Curso.objects.get(pk=curso_id)
    subjects = curso.subjects.all().order_by('nombre')
    matriculas = Matricula.objects.filter(curso=curso).select_related('estudiante')
    
    # Configuramos el ancho de las columnas (Compacto Horizontalmente, Promedio más grande)
    ws.column_dimensions['A'].width = 38 # Materias
    for c in ['B','C','D']: ws.column_dimensions[c].width = 8 # Trimesters
    ws.column_dimensions['E'].width = 16 # Promedio Final (Más grande)
    
    current_row = 1
    
    for mat in matriculas:
        # --- CABECERA POR ESTUDIANTE ---
        # Llenamos solo hasta la columna E para alinear con la tabla de abajo
        for r in range(current_row, current_row + 4):
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = FILL_LIGHT_GRAY
                
        ws.row_dimensions[current_row].height = 25
        # Mergeamos del 1 al 5 para alineación perfecta con los datos
        ws.merge_cells(start_row=current_row, end_row=current_row, start_column=1, end_column=5)
        c = ws.cell(row=current_row, column=1, value="UNIDAD EDUCATIVA \"ARQ. JULIO VITERI GAMBOA\"")
        c.font = Font(bold=True, size=14); c.alignment = CENTER_ALIGN; c.fill = FILL_LIGHT_GRAY
        
        ws.merge_cells(start_row=current_row+1, end_row=current_row+1, start_column=1, end_column=5)
        c = ws.cell(row=current_row+1, column=1, value="Educación General Bàsica Elemental")
        c.font = Font(size=11); c.alignment = CENTER_ALIGN; c.fill = FILL_LIGHT_GRAY
        
        ws.merge_cells(start_row=current_row+2, end_row=current_row+2, start_column=1, end_column=5)
        c = ws.cell(row=current_row+2, column=1, value="REPORTE ANUAL DE CALIFICACIONES")
        c.font = Font(bold=True, size=12); c.alignment = CENTER_ALIGN; c.fill = FILL_LIGHT_GRAY
        
        ws.merge_cells(start_row=current_row+3, end_row=current_row+3, start_column=1, end_column=5)
        c = ws.cell(row=current_row+3, column=1, value=f"Año Lectivo: {curso.periodo.nombre}")
        c.font = Font(bold=True, size=11); c.alignment = CENTER_ALIGN; c.fill = FILL_LIGHT_GRAY
        
        # Logos más pequeños para que entren sin problemas
        try:
            p_ind = os.path.join(settings.BASE_DIR, 'static', 'logos', 'logoExcelIndividual.png')
            p_esc = os.path.join(settings.BASE_DIR, 'static', 'logos', 'logoescuela.jpeg')
            if os.path.exists(p_ind):
                img_ind = Image(p_ind); img_ind.height = 55; img_ind.width = 130
                ws.add_image(img_ind, 'A' + str(current_row + 1))
            if os.path.exists(p_esc):
                img_esc = Image(p_esc); img_esc.height = 60; img_esc.width = 65
                ws.add_image(img_esc, 'E' + str(current_row + 1))
        except: pass
        
        r_info = current_row + 6
        ws.cell(row=r_info, column=1, value="Datos informativos del estudiante:").font=Font(size=11)
        ws.cell(row=r_info+1, column=1, value=f"Nombre: {mat.estudiante.apellidos} {mat.estudiante.nombres}").font=Font(bold=True, size=12)
        ws.cell(row=r_info+2, column=1, value=f"Grado/Curso: {curso.nivel}").font=Font(size=11)
        
        ws.cell(row=r_info+1, column=4, value="RÈGIMEN: COSTA").font=Font(size=11)
        ws.cell(row=r_info+2, column=4, value="JORNADA: MATUTINA").font=Font(size=11)
        ws.cell(row=r_info+3, column=4, value="MODALIDAD: PRESENCIAL").font=Font(size=11)
        
        # --- TABLA DE NOTAS ---
        r_table = r_info + 4
        # Headers principal
        ws.merge_cells(start_row=r_table, end_row=r_table+1, start_column=1, end_column=1)
        c = ws.cell(row=r_table, column=1, value="ASIGNATURAS:")
        c.font=Font(bold=True); c.fill=FILL_GRAY; c.alignment=CENTER_ALIGN; c.border=BORDER
        
        ws.merge_cells(start_row=r_table, end_row=r_table, start_column=2, end_column=4)
        c = ws.cell(row=r_table, column=2, value="Trimestres")
        c.font=Font(bold=True); c.fill=FILL_GRAY; c.alignment=CENTER_ALIGN; c.border=BORDER
        
        ws.merge_cells(start_row=r_table, end_row=r_table, start_column=5, end_column=5)
        c = ws.cell(row=r_table, column=5, value="Promedio Final")
        c.font=Font(bold=True); c.fill=FILL_GRAY; c.alignment=CENTER_ALIGN; c.border=BORDER
        
        # Sub Headers
        for i, v in enumerate(["I", "II", "III", "Promedio Simple"]):
            sc = ws.cell(row=r_table+1, column=2+i, value=v)
            sc.font=Font(bold=True); sc.fill=FILL_GRAY; sc.alignment=CENTER_ALIGN; sc.border=BORDER
            
        r_data = r_table + 2
        suma_promedios = 0
        for s in subjects:
            ws.cell(row=r_data, column=1, value=s.nombre.upper()).border=BORDER
            t1 = mat.get_subject_average(s, 1)
            t2 = mat.get_subject_average(s, 2)
            t3 = mat.get_subject_average(s, 3)
            pf = _trunc2((t1 + t2 + t3) / 3)
            
            suma_promedios += pf
            
            for i, v in enumerate([t1, t2, t3, pf]):
                c = ws.cell(row=r_data, column=2+i, value=v)
                c.number_format='0.00'; c.alignment=CENTER_ALIGN; c.border=BORDER
            r_data += 1
            
        # PROMEDIO GENERAL
        ws.cell(row=r_data, column=1, value="PROMEDIO GENERAL").font=Font(bold=True)
        ws.cell(row=r_data, column=1).border=BORDER
        for i in range(2, 5): ws.cell(row=r_data, column=i).border=BORDER
        pg = ws.cell(row=r_data, column=5, value=_trunc2(suma_promedios/subjects.count()) if subjects.count()>0 else 0)
        pg.font=Font(bold=True); pg.number_format='0.00'; pg.alignment=CENTER_ALIGN; pg.border=BORDER
        
        # COMPORTAMIENTO
        r_comp = r_data + 2
        ws.cell(row=r_comp, column=1, value="COMPORTAMIENTO").border=BORDER
        for i in range(1, 4):
            c_trim = Comportamiento.objects.filter(matricula=mat, trimestre=i).first()
            val = c_trim.valor if c_trim else 'B'
            sc = ws.cell(row=r_comp, column=1+i, value=val)
            sc.alignment=CENTER_ALIGN; sc.border=BORDER
            
        cf = Comportamiento.objects.filter(matricula=mat, trimestre=4).first()
        cv = ws.cell(row=r_comp, column=5, value=cf.valor if cf else 'B')
        cv.alignment=CENTER_ALIGN; cv.border=BORDER
        
        # FIRMAS
        r_firma = r_comp + 4 # 3 líneas de espacio antes de la firma
        ws.merge_cells(start_row=r_firma, end_row=r_firma, start_column=1, end_column=2)
        c = ws.cell(row=r_firma, column=1, value="_________________________")
        c.alignment=CENTER_ALIGN
        ws.merge_cells(start_row=r_firma+1, end_row=r_firma+1, start_column=1, end_column=2)
        c = ws.cell(row=r_firma+1, column=1, value="RECTOR/A")
        c.font=Font(bold=True); c.alignment=CENTER_ALIGN
        
        ws.merge_cells(start_row=r_firma, end_row=r_firma, start_column=3, end_column=4)
        c = ws.cell(row=r_firma, column=3, value="_________________________")
        c.alignment=CENTER_ALIGN
        ws.merge_cells(start_row=r_firma+1, end_row=r_firma+1, start_column=3, end_column=4)
        c = ws.cell(row=r_firma+1, column=3, value="DOCENTE TUTOR")
        c.font=Font(bold=True); c.alignment=CENTER_ALIGN
        
        current_row = r_firma + 5 # 3 líneas de separation entre actas
        
    return wb
